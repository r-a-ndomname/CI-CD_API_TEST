import json 
import csv 
from openpyxl import load_workbook


#JSON Data Provider
def read_json_data(filepath):
    with open(filepath,"r") as f:
        data_list = json.load(f)
    return [(item,) for item in data_list]

#CSV Data Provider 
def read_csv_data(filepath):
    data_list = []
    with open(filepath, "r") as f:
        # Create a CSV reader object to read each row line by line in CSV file
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            data_list.append(tuple(row))
    return data_list 

def read_excel_data(filepath):
    data_list = []
    
    # Открываем без 'with'
    workbook = load_workbook(filepath, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    
    next(rows)
    for row in rows:
        if row and row[0] is not None: 
            data_list.append(row)
            
    # Закрываем файл вручную в конце
    workbook.close() 
    return data_list
# read_csv_data(r"C:\development\Studying\API_INDUS\Day9\testdata\orders_csv_data.csv")
# read_json_data(r"C:\development\Studying\API_INDUS\Day9\testdata\orders_json_data.json")